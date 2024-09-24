from flask import Flask,render_template,request,redirect,url_for,jsonify,flash,session
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
from googleapiclient.discovery import build
from google.oauth2 import service_account
import json

app=Flask(__name__)
load_dotenv()
service_account_info = json.loads(os.getenv('GOOGLE_SERVICE_ACCOUNT_KEY'))
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
 
credentials = service_account.Credentials.from_service_account_info(service_account_info, scopes=SCOPES)
SPREADSHEET_ID = '1rEK5-TOCqydPVENrpGgaivQ6A56O1is_2eEJK2VT5IE'
service = build('sheets', 'v4', credentials=credentials)
sheet_service = service.spreadsheets()
excel_path = os.path.join(app.static_folder,'files','DMAX-2024-Live.xlsx')
workbook = load_workbook(excel_path)
sheet = workbook.active
def find_next_available_row(sheet):
    for row in range(1, sheet.max_row + 1):
        if all([cell.value in [None, ""] for cell in sheet[row]]):
            
            return row
        
    return sheet.max_row + 1 

def calculate_results(form_data,field_to_column):
    # Extract values from form_data
    values = {field:form_data.get(field, 0) for field in field_to_column.keys()}
    
    def get_value(key):
        float_fields={
            'inv_defs'
        }
        try:
            if key in float_fields:
                return float(values.get(key, 0))
            return values.get(key, 0)
        except ValueError:
            return 0
    
    # Calculate results
    

    result_AZ = sum(
        get_value(target) if get_value(actual) > 0 else 0
        for actual, target in [
            ('test_case_creation_actual', 'test_case_creation_target'),
            ('test_case_updation_actual', 'test_case_updation_target'),
            ('test_case_execution_actual', 'test_case_execution_target'),
            ('defects_found_actual', 'defects_found_target'),
            ('defects_verification_actual', 'defects_verification_target'),
            ('test_scripts_creation_actual', 'test_scripts_creation_target'),
            ('test_scripts_updation_actual', 'test_scripts_updation_target'),
            ('test_scripts_execution_actual', 'test_scripts_execution_target'),
            ('project_doc_actual', 'project_doc_target'),
            ('internal_Review_actual', 'internal_Review_target'),
            ('regression_cycle_actual', 'regression_cycle_target'),
            ('req_anal_actual', 'req_anal_target'),
            ('end_cases_exec_actual', 'end_cases_exec_target'),
            ('site_Scrub_actual', 'site_Scrub_target'),
            ('task_coverage_score_actual', 'task_coverage_score_target'),
            ('assessment_score_actual', 'assessment_score_target'),
            ('assessment_re_score_actual', 'assessment_re_score_target'),
            ('cert_score_actual', 'cert_score_target'),
            ('cert_re_score_actual', 'cert_re_score_target'),
            ('new_features_imp_actual', 'new_features_imp_target'),
            ('defects_fixed_actual', 'defects_fixed_target'),
            ('enhancements_actual', 'enhancements_target'),
            ('fig_desgns_actual', 'fig_desgns_target'),
            ('doc_update_actual', 'doc_update_target'),
            ('research_actual', 'research_target')
        ]
    )
    
    
    

    result_BA = sum(
        get_value(field) for field in [
            'test_case_creation_actual',
            'test_case_updation_actual',
            'test_case_execution_actual',
            'defects_found_actual',
            'defects_verification_actual',
            'test_scripts_creation_actual',
            'test_scripts_updation_actual',
            'test_scripts_execution_actual',
            'project_doc_actual',
            'internal_Review_actual',
            'regression_cycle_actual',
            'req_anal_actual',
            'end_cases_exec_actual',
            'site_Scrub_actual',
            'task_coverage_score_actual',
            'assessment_score_actual',
            'assessment_re_score_actual',
            'cert_score_actual',
            'cert_re_score_actual',
            'new_features_imp_actual',
            'defects_fixed_actual',
            'enhancements_actual',
            'fig_desgns_actual',
            'doc_update_actual',
            'research_actual'
        ]
    )

    result_BB = (result_BA / result_AZ * 40 / 100) if result_AZ != 0 else 0
    if get_value('client_esc') == 1:  # 'dtouch' is equivalent to BG in your field-to-column map
        result_BC=0
    else:
        result_BC = (100 - (
        get_value('inv_defs') +
        get_value('spel_errors') +
        get_value('client_esc') +
        get_value('tst_cases_missing')
        )) * 0.4 / 100

    result_BD = (get_value('att') * 1) * 10 / 100
    result_BE = get_value('dtouch') * 10 / 100 / 100
    result_BF = get_value('new_init') * 10 / 100 / 100

    result_BG = result_BB + result_BC + result_BD + result_BE + result_BF
    
    return {
        'BL': result_AZ ,
        'BM': result_BA,
        'BN': result_BB,
        'BO': result_BC,
        'BP': result_BD,
        'BQ': result_BE,
        'BR': result_BF,
        'BS': result_BG
    }

app.config['SECRET_KEY'] = 'your_secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///employees.db'
app.config['SQLALCHEMY_BINDS']={
    'dform':'sqlite:///dform.db'
}
db = SQLAlchemy(app)



class Login(db.Model):
    __tablename__ = 'login'
    __bind_key__="dform"
    id = db.Column(db.Integer, primary_key=True)
    employee_name=db.Column(db.String(100),nullable=False)
    employee_id=db.Column(db.String(100),nullable=False)
    employee_email=db.Column(db.String(100),nullable=False)
    today_date=db.Column(db.String(100),nullable=False)
    project=db.Column(db.String(100),nullable=False)
    designation=db.Column(db.String(100),nullable=False)
    test_case_creation_target= db.Column(db.Integer)
    test_case_creation_actual=db.Column(db.Integer)
    test_case_updation_target=db.Column(db.Integer)
    test_case_updation_actual=db.Column(db.Integer)
    test_case_execution_target=db.Column(db.Integer)
    test_case_execution_actual=db.Column(db.Integer)
    defects_found_target=db.Column(db.Integer)
    defects_found_actual=db.Column(db.Integer)
    test_scripts_creation_target=db.Column(db.Integer)
    test_scripts_creation_actual=db.Column(db.Integer)
    test_scripts_updation_target=db.Column(db.Integer)
    test_scripts_updation_actual=db.Column(db.Integer)
    test_scripts_execution_target=db.Column(db.Integer)
    test_scripts_execution_actual=db.Column(db.Integer)
    site_Scrub_target=db.Column(db.Integer)
    site_Scrub_actual=db.Column(db.Integer)
    project_doc_target=db.Column(db.Integer)
    project_doc_actual=db.Column(db.Integer)
    internal_Review_target=db.Column(db.Integer)
    internal_Review_actual=db.Column(db.Integer)
    regression_cycle_target=db.Column(db.Integer)
    regression_cycle_actual=db.Column(db.Integer)
    req_anal_target=db.Column(db.Integer)
    req_anal_actual=db.Column(db.Integer)
    end_cases_exec_target=db.Column(db.Integer)
    end_cases_exec_actual=db.Column(db.Integer)
    task_coverage_score_target=db.Column(db.Integer)
    task_coverage_score_actual=db.Column(db.Integer)
    assessment_score_target=db.Column(db.Integer)
    assessment_score_actual=db.Column(db.Integer)
    assessment_re_score_target=db.Column(db.Integer)
    assessment_re_score_actual=db.Column(db.Integer)
    cert_score_target=db.Column(db.Integer)
    cert_score_actual=db.Column(db.Integer)
    cert_re_score_target=db.Column(db.Integer)
    cert_re_score_actual=db.Column(db.Integer)
    new_features_imp_target=db.Column(db.Integer)
    new_features_imp_actual=db.Column(db.Integer)
    defects_fixed_target=db.Column(db.Integer)
    defects_fixed_actual=db.Column(db.Integer)
    enhancements_target=db.Column(db.Integer)
    enhancements_actual=db.Column(db.Integer)
    fig_desgns_target=db.Column(db.Integer)
    fig_desgns_actual=db.Column(db.Integer)
    doc_update_target=db.Column(db.Integer)
    doc_update_actual=db.Column(db.Integer)
    research_target=db.Column(db.Integer)
    research_actual=db.Column(db.Integer)
    inv_defs=db.Column(db.Integer)
    spel_errors=db.Column(db.Float)
    client_esc=db.Column(db.Integer)
    tst_cases_missing=db.Column(db.Integer)
    att=db.Column(db.Integer)
    dtouch=db.Column(db.Integer)
    new_init=db.Column(db.Integer)
    defects_verification_target=db.Column(db.Integer)
    defects_verification_actual=db.Column(db.Integer)

# Employee Model
class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    emp_id = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(100), nullable=False)

def fetch_and_print_sheet_values():
    range_name = 'Sheet1!A1:Z'  # Adjust the range as needed
    result = sheet_service.values().get(spreadsheetId=SPREADSHEET_ID, range=range_name).execute()
    values = result.get('values', [])
    
    if not values:
        print('No data found.')
    else:
        for row in values:
            pass

def append_sheet_values(values):
    range_name = 'Sheet1!A1:Z'
    
    body={"values":[values]}
    result=(
        service.spreadsheets().values()
        .append(
            spreadsheetId=SPREADSHEET_ID, 
            range=range_name,
            valueInputOption='RAW',
            body=body
        ).execute()
    )
    print(result,"got appended")

@app.route('/form',methods=["GET","POST"])
def home():
    if 'username' in session:
        username = session['username']
        employee = Employee.query.filter_by(emp_id=username).first()
        if employee:
            role = employee.role
            print(role)
            
    else:
        return redirect(url_for('sign'))

    if request.method=="POST":
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        
        next_row = find_next_available_row(sheet)
        
        field_to_column = {
            "employee_name": 'A',
            "employee_id": 'B',
            "employee_email": 'C',
            "today_date": 'D',
            "project": 'E',
            "designation": 'F',
            "test_case_creation_target": 'G',
            "test_case_creation_actual": 'H',
            "test_case_updation_target": 'I',
            "test_case_updation_actual": 'J',
            "test_case_execution_target": 'K',
            "test_case_execution_actual": 'L',
            "defects_found_target":'M',
            "defects_found_actual":'N',
            "defects_verification_target":'O',
            "defects_verification_actual":'P',
            "test_scripts_creation_target":'Q',
            "test_scripts_creation_actual":'R',
            "test_scripts_updation_target":'S',
            "test_scripts_updation_actual":'T',
            "test_scripts_execution_target":'U',
            "test_scripts_execution_actual":'V',
            "site_Scrub_target":'AG',
            "site_Scrub_actual":'AH',
            "project_doc_target":'W',
            "project_doc_actual":'X',
            "internal_Review_target":'Y',
            "internal_Review_actual":'Z',
            "regression_cycle_target":'AA',
            "regression_cycle_actual":'AB',
            "req_anal_target":'AC',
            "req_anal_actual":'AD',
            "end_cases_exec_target":'AE',
            "end_cases_exec_actual":'AF',
            "task_coverage_score_target":'AI',
            "task_coverage_score_actual":'AJ',
            "assessment_score_target":'AK',
            "assessment_score_actual":'AL',
            "assessment_re_score_target":'AM',
            "assessment_re_score_actual":'AN',
            "cert_score_target":"AO",
            "cert_score_actual":'AP',
            "cert_re_score_target":'AQ',
            "cert_re_score_actual":'AR',
            "new_features_imp_target":'AS',
            "new_features_imp_actual":'AT',
            "defects_fixed_target":'AU',
            "defects_fixed_actual":'AV',
            "enhancements_target":'AW',
            "enhancements_actual":'AX',
            "fig_desgns_target":'AY',
            "fig_desgns_actual":'AZ',
            "doc_update_target":'BA',
            "doc_update_actual":'BB',
            "research_target":'BC',
            "research_actual":'BD',
            "inv_defs":'BE',
            "spel_errors":'BF',
            "client_esc":'BG',
            "tst_cases_missing":'BH',
            "att":'BI',
            "dtouch":'BJ',
            "new_init":'BK',    
        }
        form_data = {}
        row_values = []
        for field, column in field_to_column.items():
            value = request.form.get(field)
            value = value.strip() if value else ''
            if value and value.replace('.', '', 1).isdigit():
                value = float(value)
            form_data[field] = value
            sheet[f'{column}{next_row}'] = value
            row_values.append(value)
          

        formulas = {
            'BL': f"=SUM(IF(H{next_row}>0,G{next_row}),IF(J{next_row}>0,I{next_row}),IF(L{next_row}>0,K{next_row}),IF(N{next_row}>0,M{next_row}),IF(P{next_row}>0,O{next_row}),IF(R{next_row}>0,Q{next_row}),IF(T{next_row}>0,S{next_row}),IF(V{next_row}>0,U{next_row}),IF(X{next_row}>0,W{next_row}),IF(Z{next_row}>0,Y{next_row}),IF(AB{next_row}>0,AA{next_row}),IF(AD{next_row}>0,AC{next_row}),IF(AF{next_row}>0,AE{next_row}),IF(AH{next_row}>0,AG{next_row}),IF(AJ{next_row}>0,AI{next_row}),IF(AL{next_row}>0,AK{next_row}),IF(AN{next_row}>0,AM{next_row}),IF(AP{next_row}>0,AO{next_row}),IF(AR{next_row}>0,AQ{next_row}))",
            'BM': f"=SUM(H{next_row},J{next_row},L{next_row},N{next_row},P{next_row},R{next_row},T{next_row},V{next_row},X{next_row},Z{next_row},AB{next_row},AD{next_row},AF{next_row},AH{next_row},AJ{next_row},AL{next_row},AN{next_row},AP{next_row},AR{next_row})",
            'BN': f"=BA{next_row}/AZ{next_row}*40/100",
            'BO': f"=IF(AU{next_row}=1,0,100-(SUM($AS{next_row}:$AV{next_row})))*0.4/100",
            'BP': f"=(AW{next_row}*1)*10/100",
            'BQ': f"=AX{next_row}*10/100/100",
            'BR': f"=AY{next_row}*10/100/100",
            'BS': f"=SUM(BB{next_row}:BF{next_row})"
        }
        calculated_results = calculate_results(form_data,field_to_column)
        for column, formula in formulas.items():
            sheet[f'{column}{next_row}'] = formula
        for row in range(next_row + 1, sheet.max_row + 1):
            for col in field_to_column.values():
                sheet[f'{col}{row}'] = None    
        workbook.save(excel_path)
        append_data = {
            "values": [row_values + [calculated_results[col] for col in formulas.keys()]]
        }
        sheet_service.values().append(
            spreadsheetId=SPREADSHEET_ID,
            range='Sheet1!A:AY',  # Adjust to your sheet and range
            valueInputOption='USER_ENTERED',  # This allows Google Sheets to interpret values/formulas
            body=append_data
        ).execute()

        new_entry = Login(
            employee_name=form_data['employee_name'],
            employee_id=form_data['employee_id'],
            employee_email=form_data['employee_email'],
            today_date=form_data['today_date'],
            project=form_data['project'],
            designation=form_data['designation'],
            test_case_creation_target=form_data.get('test_case_creation_target'),
            test_case_creation_actual=form_data.get('test_case_creation_actual'),
            test_case_updation_target=form_data.get('test_case_updation_target'),
            test_case_updation_actual=form_data.get('test_case_updation_actual'),
            test_case_execution_target=form_data.get('test_case_execution_target'),
            test_case_execution_actual=form_data.get('test_case_execution_actual'),
            defects_found_target=form_data.get('defects_found_target'),
            defects_found_actual=form_data.get('defects_found_actual'),
            test_scripts_creation_target=form_data.get('test_scripts_creation_target'),
            test_scripts_creation_actual=form_data.get('test_scripts_creation_actual'),
            test_scripts_updation_target=form_data.get('test_scripts_updation_target'),
            test_scripts_updation_actual=form_data.get('test_scripts_updation_actual'),
            test_scripts_execution_target=form_data.get('test_scripts_execution_target'),
            test_scripts_execution_actual=form_data.get('test_scripts_execution_actual'),
            site_Scrub_target=form_data.get('site_Scrub_target'),
            site_Scrub_actual=form_data.get('site_Scrub_actual'),
            project_doc_target=form_data.get('project_doc_target'),
            project_doc_actual=form_data.get('project_doc_actual'),
            internal_Review_target=form_data.get('internal_Review_target'),
            internal_Review_actual=form_data.get('internal_Review_actual'),
            regression_cycle_target=form_data.get('regression_cycle_target'),
            regression_cycle_actual=form_data.get('regression_cycle_actual'),
            req_anal_target=form_data.get('req_anal_target'),
            req_anal_actual=form_data.get('req_anal_actual'),
            end_cases_exec_target=form_data.get('end_cases_exec_target'),
            end_cases_exec_actual=form_data.get('end_cases_exec_actual'),
            task_coverage_score_target=form_data.get('task_coverage_score_target'),
            task_coverage_score_actual=form_data.get('task_coverage_score_actual'),
            assessment_score_target=form_data.get('assessment_score_target'),
            assessment_score_actual=form_data.get('assessment_score_actual'),
            assessment_re_score_target=form_data.get('assessment_re_score_target'),
            assessment_re_score_actual=form_data.get('assessment_re_score_actual'),
            cert_score_target=form_data.get('cert_score_target'),
            cert_score_actual=form_data.get('cert_score_actual'),
            cert_re_score_target=form_data.get('cert_re_score_target'),
            cert_re_score_actual=form_data.get('cert_re_score_actual'),
            new_features_imp_target=form_data.get('new_features_imp_target'),
            new_features_imp_actual=form_data.get('new_features_imp_actual'),
            defects_fixed_target=form_data.get('defects_fixed_target'),
            defects_fixed_actual=form_data.get('defects_fixed_actual'),
            defects_verification_target=form_data.get('defects_verification_target'),
            defects_verification_actual=form_data.get('defects_verification_actual'),
            enhancements_target=form_data.get('enhancements_target'),
            enhancements_actual=form_data.get('enhancements_actual'),
            fig_desgns_target=form_data.get('fig_desgns_target'),
            fig_desgns_actual=form_data.get('fig_desgns_actual'),
            doc_update_target=form_data.get('doc_update_target'),
            doc_update_actual=form_data.get('doc_update_actual'),
            research_target=form_data.get('research_target'),
            research_actual=form_data.get('research_actual'),
            inv_defs=form_data.get('inv_defs'),
            spel_errors=form_data.get('spel_errors'),
            client_esc=form_data.get('client_esc'),
            tst_cases_missing=form_data.get('tst_cases_missing'),
            att=form_data.get('att'),
            dtouch=form_data.get('dtouch'),
            new_init=form_data.get('new_init')
        )

        # Add to DB and commit the session
        db.session.add(new_entry)
        db.session.commit()
        
        
        return redirect(url_for('home'))
    return render_template('index.html',role=role)

@app.route('/print_sheet_values', methods=["GET"])
def print_sheet_values():
    values=[10,20,30,40,50]
    append_sheet_values(values)
    fetch_and_print_sheet_values()
    return "Check your console for the printed values."    
    
@app.route('/',methods=["GET","POST"])
def sign():
    if request.method=="POST":
        username = request.form["username"]  
        password = request.form["password"]
        employee = Employee.query.filter_by(emp_id=username).first()
        if employee and employee.password == password:
            session['username'] = username
            
            return redirect(url_for('home'))
        else:
            print("invalid credentials")
        
        
    return render_template('sign.html')     

    
@app.route('/read_excel')
def read_excel():
    # Construct the path to the Excel file in the static folder
    

    # Load the Excel workbook
    
    
    # Select the active sheet
    sheet = workbook.active
    
    # Example: Reading data from the first row, first column (A1)
    first_cell_value = sheet['A3'].value

    # Optionally: Process the data further and return it to the template
    return f"Value in A1: {first_cell_value}"


@app.route('/search', methods=['POST'])
def search_employee():
    data = request.json
    employee_name = data.get('employee_name')
    
    # Query the Login table where employee_name matches (case-insensitive search)
    matched_employees = Login.query.filter(Login.employee_name.ilike(f'%{employee_name}%')).all()

    # Create a list of dictionaries containing employee details
    employees_list = [
        {column.name: getattr(emp, column.name) for column in Login.__table__.columns}
        for emp in matched_employees
    ]
    
    return jsonify({"employees": employees_list})

@app.route('/register', methods=['GET','POST'])
def register():
    if request.method=="POST":
        username=request.form["username"]
        password=request.form["password"]
        role=request.form["role"]
        existing_employee = Employee.query.filter_by(emp_id=username).first()
        if existing_employee:
            
            flash('Employee with that ID already exists', 'error')
            return redirect('/register')
        new_employee = Employee(emp_id=username, password=password, role=role)
        db.session.add(new_employee)
        db.session.commit()
        flash('Employee registered successfully!', 'success')
        return redirect('/')
    return render_template("register.html")

@app.route("/no-acess")
def no_access():
    return render_template("no_acess.html")

@app.route("/logout")
def logout():
    session.pop('username',None)
    return redirect(url_for('sign'))

with app.app_context():
        
        db.create_all()
        

if __name__ == "__main__":
    app.run(debug=True)